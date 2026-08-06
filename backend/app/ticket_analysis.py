import csv
import json
import math
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook

from .config import configured_model
from .diagnostics import diagnostic_event
from .llm import _chat, _json_object
from .ticket_taxonomy import DEFAULT_TAXONOMY, TaxonomyClassification, TaxonomyRule, classify_ticket_v2, find_taxonomy_match, match_taxonomy_rule, normalize_signal


TITLE_FIELDS = ("short_description", "short description", "title", "summary")
DESCRIPTION_FIELDS = ("description", "details")
ID_FIELDS = ("ticket_id", "ticket id", "number", "incident_number", "incident number", "incident_no", "incident no", "sys_id", "sys id")
STRONG_ID_FIELDS = ("number", "sys_id", "sys id", "ticket_id", "ticket id", "incident_number", "incident number", "incident_no", "incident no")
METADATA_FIELDS = {
    "category", "subcategory", "assignment_group", "assignment group", "priority",
    "severity", "state", "created_at", "created at", "resolved_at", "resolved at",
    "application", "service", "ci", "record_type", "record type", "type",
    "source_system", "source system", "business_service", "business service",
    "cmdb_ci", "cmdb ci", "component", "labels", "comments", "work_notes", "work notes",
}
STOP_WORDS = {
    "a", "an", "and", "are", "as", "at", "be", "been", "but", "by", "can",
    "could", "for", "from", "has", "have", "in", "is", "it", "of", "on", "or",
    "our", "that", "the", "their", "this", "to", "unable", "user", "users", "was",
    "were", "with", "after", "before", "issue", "issues", "incident", "incidents",
    "problem", "ticket", "tickets", "title", "description", "reports", "reported",
    "initial", "triage", "captured", "common",
}
LLM_FALLBACK_MAX_ATOMICS = 40
LLM_FALLBACK_MIN_CONFIDENCE = 0.58
LLM_NAMING_MAX_GROUPS = 24
LLM_NAMING_MAX_EXAMPLES = 4
# Renaming this bucket would break the unresolved-ticket accounting in the
# pipeline trace and the overflow rules below, which both match it by name.
LLM_NAMING_PROTECTED_NAMES = {"other service issues"}



def _key(value: Any) -> str:
    return re.sub(r"[_\s-]+", " ", str(value or "").strip().lower())


def _text(value: Any) -> str | None:
    if value is None:
        return None
    clean = re.sub(r"\s+", " ", str(value)).strip()
    return clean or None


@dataclass
class NormalizedTicket:
    ticket_id: str | None
    title: str | None
    description: str | None
    primary_text: str
    metadata: dict[str, Any] = field(default_factory=dict)


def normalize_ticket(row: dict[str, Any]) -> NormalizedTicket:
    indexed = {_key(name): value for name, value in row.items()}

    def first(aliases: Iterable[str]) -> str | None:
        return next((_text(indexed.get(_key(alias))) for alias in aliases if _text(indexed.get(_key(alias)))), None)

    title = first(TITLE_FIELDS)
    description = first(DESCRIPTION_FIELDS)
    ticket_id = first(ID_FIELDS)
    if title and description:
        primary_text = f"{title}\n{description}"
    else:
        primary_text = title or description or ""
    if not primary_text:
        excluded = {_key(value) for value in (*ID_FIELDS, *METADATA_FIELDS)}
        fallback = [_text(value) for name, value in indexed.items() if name not in excluded and _text(value)]
        primary_text = " | ".join(fallback)
        title = fallback[0] if fallback else None
    metadata = {
        name: value for name, value in row.items()
        if _key(name) in {_key(item) for item in METADATA_FIELDS} and _text(value)
    }
    return NormalizedTicket(ticket_id, title, description, primary_text, metadata)


def read_ticket_rows(path: Path) -> list[dict[str, Any]]:
    extension = path.suffix.lower()
    if extension in {".csv", ".tsv"}:
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            return list(csv.DictReader(handle, delimiter="\t" if extension == ".tsv" else ","))
    if extension in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        rows: list[dict[str, Any]] = []
        try:
            for sheet in workbook.worksheets:
                values = sheet.iter_rows(values_only=True)
                headers = next(values, None)
                if not headers:
                    continue
                names = [str(value or f"Column {index + 1}") for index, value in enumerate(headers)]
                rows.extend(dict(zip(names, row)) for row in values)
        finally:
            workbook.close()
        return rows
    if extension == ".json":
        data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        if isinstance(data, dict):
            for key in ("tickets", "incidents", "records", "data", "items"):
                if isinstance(data.get(key), list):
                    data = data[key]
                    break
            else:
                data = [data]
        if not isinstance(data, list):
            raise ValueError("JSON ticket data must be an array or contain a records array")
        return [item if isinstance(item, dict) else {"description": item} for item in data]
    if extension in {".txt", ".md"}:
        lines = [line.strip() for line in path.read_text(encoding="utf-8", errors="replace").splitlines() if line.strip()]
        return [{"description": line} for line in lines]
    raise ValueError("Ticket Analysis supports CSV, TSV, XLSX, XLSM, JSON, TXT, and MD files")


def _normalized_row_fingerprint(row: dict[str, Any]) -> str:
    values = [f"{_key(name)}={_text(value).casefold()}" for name, value in sorted(row.items(), key=lambda item: _key(item[0])) if _text(value)]
    return "|".join(values)


def clean_tickets(rows: list[dict[str, Any]]) -> tuple[list[NormalizedTicket], int, int]:
    tickets: list[NormalizedTicket] = []
    seen_identifiers: set[tuple[str, str]] = set()
    seen_rows: set[str] = set()
    empty = duplicates = 0
    for row in rows:
        ticket = normalize_ticket(row)
        indexed = {_key(name): _text(value) for name, value in row.items()}
        identifiers = [(_key(field), indexed[_key(field)].casefold()) for field in STRONG_ID_FIELDS if indexed.get(_key(field))]
        row_fingerprint = _normalized_row_fingerprint(row)
        if not ticket.primary_text.strip():
            empty += 1
        elif (identifiers and any(identifier in seen_identifiers for identifier in identifiers)) or (not identifiers and row_fingerprint in seen_rows):
            duplicates += 1
        else:
            seen_identifiers.update(identifiers)
            if not identifiers:
                seen_rows.add(row_fingerprint)
            tickets.append(ticket)
    return tickets, empty, duplicates


def _tokens(text: str) -> Counter[str]:
    words = re.findall(r"[a-z][a-z0-9]{2,}", text.lower())
    return Counter(word for word in words if word not in STOP_WORDS)


def _cosine(left: Counter[str], right: Counter[str]) -> float:
    shared = left.keys() & right.keys()
    numerator = sum(left[word] * right[word] for word in shared)
    denominator = math.sqrt(sum(value * value for value in left.values()) * sum(value * value for value in right.values()))
    return numerator / denominator if denominator else 0.0


def _idf(token_sets: list[Counter[str]]) -> dict[str, float]:
    documents = len(token_sets) or 1
    frequency = Counter(token for tokens in token_sets for token in tokens)
    return {token: math.log((documents + 1) / (count + 1)) + 1.0 for token, count in frequency.items()}


def _dense_cosine(left: list[float], right: list[float]) -> float:
    numerator = sum(a * b for a, b in zip(left, right))
    denominator = math.sqrt(sum(a * a for a in left) * sum(b * b for b in right))
    return numerator / denominator if denominator else 0.0


def _build_vectors(texts: list[str], method: str) -> tuple[list[Any], Callable[[Any, Any], float]]:
    """Vectorize signature texts with the embedding backend the run selected.

    `tfidf` is genuinely inverse-document-frequency weighted rather than raw term
    counts, so boilerplate shared by every ticket stops dominating similarity.
    `neural_hash` reuses the app's embedding stack (fastembed, hash fallback);
    `hybrid` averages both so lexical overlap and semantic proximity must agree.
    """
    token_sets = [_tokens(text) for text in texts]
    if method == "tfidf":
        weights = _idf(token_sets)
        sparse = [Counter({token: count * weights.get(token, 1.0) for token, count in tokens.items()}) for tokens in token_sets]
        return sparse, _cosine
    if method == "neural_hash":
        from .vector_store import embed_passages
        return embed_passages(texts), _dense_cosine
    if method == "hybrid":
        from .vector_store import embed_passages
        weights = _idf(token_sets)
        sparse = [Counter({token: count * weights.get(token, 1.0) for token, count in tokens.items()}) for tokens in token_sets]
        dense = embed_passages(texts)
        return list(zip(sparse, dense)), lambda left, right: (_cosine(left[0], right[0]) + _dense_cosine(left[1], right[1])) / 2
    return token_sets, _cosine


def _similarity_matrix(vectors: list[Any], similarity: Callable[[Any, Any], float]) -> list[list[float]]:
    size = len(vectors)
    matrix = [[1.0] * size for _ in range(size)]
    for left in range(size):
        for right in range(left + 1, size):
            score = similarity(vectors[left], vectors[right])
            matrix[left][right] = matrix[right][left] = score
    return matrix


def _connected_components(size: int, edges: Iterable[tuple[int, int]]) -> list[list[int]]:
    parents = list(range(size))

    def find(value: int) -> int:
        while parents[value] != value:
            parents[value] = parents[parents[value]]
            value = parents[value]
        return value

    for left, right in edges:
        left_root, right_root = find(left), find(right)
        if left_root != right_root:
            parents[right_root] = left_root
    grouped: dict[int, list[int]] = {}
    for index in range(size):
        grouped.setdefault(find(index), []).append(index)
    return list(grouped.values())


def _agglomerative_clusters(matrix: list[list[float]], threshold: float, target_clusters: int | None) -> list[list[int]]:
    """Average-link agglomeration: merge the closest pair until the link falls
    below the threshold, or until the requested cluster count is reached."""
    clusters = [[index] for index in range(len(matrix))]
    floor = max(1, target_clusters or 1)
    while len(clusters) > floor:
        best = (0.0, -1, -1)
        for left in range(len(clusters)):
            for right in range(left + 1, len(clusters)):
                pairs = [matrix[a][b] for a in clusters[left] for b in clusters[right]]
                score = sum(pairs) / len(pairs)
                if score > best[0]:
                    best = (score, left, right)
        score, left, right = best
        if left < 0 or (score < threshold and len(clusters) <= (target_clusters or 0)):
            break
        if score < threshold and not target_clusters:
            break
        if score <= 0:
            break
        clusters[left].extend(clusters.pop(right))
    return clusters


def _kmeans_clusters(vectors: list[Any], similarity: Callable[[Any, Any], float], target_clusters: int | None) -> list[list[int]]:
    """k-means in similarity space: seeded farthest-first, then assign-to-nearest
    -medoid rounds. Medoids keep this valid for sparse Counters and dense lists
    alike, which a mean-based centroid would not be."""
    size = len(vectors)
    k = max(1, min(target_clusters or max(1, int(math.sqrt(size))), size))
    medoids = [0]
    while len(medoids) < k:
        farthest, best_distance = None, -1.0
        for index in range(size):
            if index in medoids:
                continue
            distance = min(1.0 - similarity(vectors[index], vectors[medoid]) for medoid in medoids)
            if distance > best_distance:
                farthest, best_distance = index, distance
        if farthest is None:
            break
        medoids.append(farthest)
    assignments = [0] * size
    for _ in range(8):
        clusters: dict[int, list[int]] = {medoid: [] for medoid in medoids}
        for index in range(size):
            best_medoid = max(medoids, key=lambda medoid: similarity(vectors[index], vectors[medoid]))
            clusters[best_medoid].append(index)
            assignments[index] = best_medoid
        updated = []
        for medoid, members in clusters.items():
            if not members:
                continue
            best = max(members, key=lambda candidate: sum(similarity(vectors[candidate], vectors[other]) for other in members))
            updated.append(best)
        if sorted(updated) == sorted(medoids):
            break
        medoids = updated or medoids
    grouped: dict[int, list[int]] = {}
    for index, medoid in enumerate(assignments):
        grouped.setdefault(medoid, []).append(index)
    return [members for members in grouped.values() if members]


def _hdbscan_lite_clusters(matrix: list[list[float]], threshold: float, min_samples: int) -> tuple[list[list[int]], list[int]]:
    """Density-based grouping: only points with enough close neighbours are core
    points and may extend a cluster. Everything else is returned as noise instead
    of being forced into a group."""
    size = len(matrix)
    neighbours = [
        [other for other in range(size) if other != index and matrix[index][other] >= threshold]
        for index in range(size)
    ]
    core = {index for index in range(size) if len(neighbours[index]) >= max(1, min_samples) - 1}
    edges = [
        (index, other)
        for index in core
        for other in neighbours[index]
        if other in core or len(neighbours[other]) >= 1
    ]
    components = _connected_components(size, edges)
    clusters, noise = [], []
    for component in components:
        if any(index in core for index in component) and len(component) > 1:
            clusters.append(component)
        else:
            noise.extend(component)
    return clusters, noise


def _kwikbucks_clusters(matrix: list[list[float]], threshold: float, oracle_budget: int) -> list[list[int]]:
    """KwikBucks-style correlation clustering: the cheap similarity signal ranks
    candidate pairs, and a limited budget of stricter 'oracle' comparisons (a
    higher bar on the same matrix) decides which edges are real. Pairs the budget
    never reaches stay unlinked, so precision beats recall here by design."""
    size = len(matrix)
    ranked = sorted(
        ((matrix[left][right], left, right) for left in range(size) for right in range(left + 1, size)),
        reverse=True,
    )
    oracle_threshold = min(0.95, threshold * 1.25)
    edges = []
    for score, left, right in ranked[:max(1, oracle_budget)]:
        if score >= threshold and score >= oracle_threshold:
            edges.append((left, right))
    return _connected_components(size, edges)


def _calibrated_threshold(matrix: list[list[float]], threshold: float, base: float, span: float) -> float:
    """Turn the similarity slider into a percentile of the observed distribution.

    Absolute cosine is not comparable across vector spaces: lexical vectors are
    sparse and mostly near-orthogonal, while dense embeddings put nearly every
    pair above 0.7, so a single fixed cut either merges everything or nothing.
    Reading the slider as "how selective to be" keeps it meaningful in both.

    `base`/`span` set how selective the calling method needs to be. Single-link
    methods chain — one surviving edge welds two clusters together — so they ask
    for a far higher percentile than average-link or medoid methods, which are
    not chaining-prone.
    """
    scores = sorted(matrix[left][right] for left in range(len(matrix)) for right in range(left + 1, len(matrix)))
    if not scores:
        return threshold
    percentile = min(0.995, max(0.05, base + span * threshold))
    return scores[min(len(scores) - 1, int(percentile * len(scores)))]


def _discovery_clusters(
    texts: list[str],
    *,
    embedding_method: str,
    clustering_method: str,
    threshold: float,
    target_clusters: int | None,
    min_samples: int,
) -> tuple[list[list[int]], list[int]]:
    """Route signature texts through the selected embedding + clustering pair.

    Returns (clusters, noise). Only hdbscan_lite produces real noise; the other
    methods place every signature, and small or incoherent clusters are filtered
    downstream by min_group_size instead.
    """
    if not texts:
        return [], []
    vectors, similarity = _build_vectors(texts, embedding_method)
    if clustering_method == "kmeans":
        return _kmeans_clusters(vectors, similarity, target_clusters), []
    matrix = _similarity_matrix(vectors, similarity)
    # Dense spaces, and density clustering in any space, need a distribution-
    # relative radius; the lexical default keeps its absolute cosine cut.
    dense = embedding_method in {"neural_hash", "hybrid"}
    if dense and clustering_method in {"taxonomy_semantic", "google_kwikbucks"}:
        threshold = _calibrated_threshold(matrix, threshold, 0.90, 0.09)
    elif clustering_method == "hdbscan_lite":
        threshold = _calibrated_threshold(matrix, threshold, 0.92, 0.07)
    elif dense:
        threshold = _calibrated_threshold(matrix, threshold, 0.40, 0.60)
    if clustering_method == "agglomerative":
        return _agglomerative_clusters(matrix, threshold, target_clusters), []
    if clustering_method == "hdbscan_lite":
        return _hdbscan_lite_clusters(matrix, threshold, min_samples)
    if clustering_method == "google_kwikbucks":
        budget = max(len(texts) * 4, (target_clusters or 12) * 8)
        return _kwikbucks_clusters(matrix, threshold, budget), []
    edges = [
        (left, right)
        for left in range(len(texts))
        for right in range(left + 1, len(texts))
        if matrix[left][right] >= threshold
    ]
    return _connected_components(len(texts), edges), []


def _group_label(tickets: list[NormalizedTicket]) -> tuple[str, str, float]:
    token_sets = [_tokens(ticket.primary_text) for ticket in tickets]
    document_frequency = Counter(token for tokens in token_sets for token in tokens)
    common = [token for token, _ in document_frequency.most_common(4)]
    seed = tickets[0].title or " ".join(common[:3])
    cleaned_seed = " ".join(word for word in re.findall(r"[A-Za-z][A-Za-z0-9-]*", seed) if word.lower() not in STOP_WORDS)
    label = (cleaned_seed or "Unclassified service").strip().title()
    if not re.search(r"\b(issue|failure|error|problem|alert)s?$", label, re.I):
        label += " Issues"
    examples = [ticket.title or ticket.description or ticket.primary_text for ticket in tickets[:2]]
    description = f"Users are affected by related problems such as {_human_list(examples)}."
    confidence = min(0.99, 0.55 + (0.1 * min(len(tickets), 3)) + (0.08 if common else 0))
    return label, description[:500], round(confidence, 2)


def _metadata(ticket: NormalizedTicket, field: str) -> str | None:
    wanted = _key(field)
    return next((_text(value) for name, value in ticket.metadata.items() if _key(name) == wanted and _text(value)), None)


def _human_list(values: Iterable[str]) -> str:
    clean = list(dict.fromkeys(value.strip().rstrip(".") for value in values if value and value.strip()))
    if not clean:
        return "service interruptions"
    if len(clean) == 1:
        return clean[0]
    if len(clean) == 2:
        return f"{clean[0]} and {clean[1]}"
    return ", ".join(clean[:-1]) + f", and {clean[-1]}"


@dataclass
class _AtomicGroup:
    tickets: list[NormalizedTicket]
    category: str | None
    subcategory: str | None


def _using_default_taxonomy(taxonomy: tuple[TaxonomyRule, ...]) -> bool:
    return taxonomy is DEFAULT_TAXONOMY or (
        len(taxonomy) == len(DEFAULT_TAXONOMY)
        and all(left.name == right.name for left, right in zip(taxonomy, DEFAULT_TAXONOMY))
    )


def _v2_match_atomic(atomic: _AtomicGroup) -> TaxonomyClassification | None:
    classifications = [classify_ticket_v2(ticket) for ticket in atomic.tickets]
    eligible = [
        classification for classification in classifications
        if classification.category != "Unclassified / Needs Review"
        and (classification.confidence in {"high", "medium"} or classification.score >= 4.5)
    ]
    if not eligible:
        return None
    grouped: dict[str, list[TaxonomyClassification]] = {}
    for classification in eligible:
        grouped.setdefault(classification.category, []).append(classification)
    category, matches = max(
        grouped.items(),
        key=lambda item: (
            len(item[1]),
            sum(match.score for match in item[1]) / max(1, len(item[1])),
        ),
    )
    best = max(matches, key=lambda match: match.score)
    if best.confidence == "low" and len(matches) < len(atomic.tickets) * 0.6:
        return None
    return best


def _initial_groups(tickets: list[NormalizedTicket]) -> list[_AtomicGroup]:
    """Create narrow groups first; taxonomy rollup happens in a separate pass."""
    buckets: dict[tuple[str, str, str], list[NormalizedTicket]] = {}
    labels: dict[tuple[str, str, str], tuple[str | None, str | None]] = {}
    for ticket in tickets:
        category = _metadata(ticket, "category")
        subcategory = _metadata(ticket, "subcategory")
        if category or subcategory:
            key = ("structured", _key(category), _key(subcategory))
        else:
            # Exact normalized issue text is deliberately narrow. Related titles
            # are merged by the parent taxonomy rather than exposed to users.
            seed = ticket.title or ticket.description or ticket.primary_text
            key = ("text", _key(seed), "")
        buckets.setdefault(key, []).append(ticket)
        labels.setdefault(key, (category, subcategory))
    return [_AtomicGroup(members, *labels[key]) for key, members in buckets.items()]


def _representatives(atomic_groups: list[_AtomicGroup], count: int) -> list[dict[str, Any]]:
    examples = []
    seen_titles: set[str] = set()
    for atomic in sorted(atomic_groups, key=lambda item: len(item.tickets), reverse=True):
        for ticket in atomic.tickets:
            title = ticket.title or ticket.description or ticket.primary_text
            if _key(title) in seen_titles:
                continue
            examples.append({"ticketId": ticket.ticket_id, "title": title})
            seen_titles.add(_key(title))
            break
        if len(examples) >= count:
            break
    return examples


def _other_service_group(
    atomics: list[_AtomicGroup],
    representative_count: int,
    reason: str = "fallback: tiny or low-confidence semantic clusters",
) -> dict[str, Any]:
    return {
        "groupName": "Other Service Issues",
        "description": "Small or low-confidence patterns that could not yet form a reliable business pain point.",
        "incidentCount": sum(len(atomic.tickets) for atomic in atomics),
        "representativeTickets": _representatives(atomics, representative_count),
        "confidence": 0.4,
        "matched_reason": reason,
    }


def _hierarchical_groups(
    tickets: list[NormalizedTicket],
    representative_count: int,
    taxonomy: tuple[TaxonomyRule, ...],
) -> tuple[list[dict[str, Any]], list[_AtomicGroup]]:
    parent_buckets: dict[str, dict[str, Any]] = {}
    fallback_buckets: dict[str, list[_AtomicGroup]] = {}
    unclassified: list[_AtomicGroup] = []
    use_v2 = _using_default_taxonomy(taxonomy)
    for atomic in _initial_groups(tickets):
        v2_match = _v2_match_atomic(atomic) if use_v2 else None
        if v2_match and v2_match.rule:
            bucket = parent_buckets.setdefault(v2_match.rule.name, {"rule": v2_match.rule, "atomics": [], "matches": [], "v2": True})
            bucket["atomics"].append(atomic)
            bucket["matches"].append(v2_match)
        else:
            match = find_taxonomy_match(
                atomic.category,
                atomic.subcategory,
                [ticket.title or "" for ticket in atomic.tickets],
                [ticket.description or "" for ticket in atomic.tickets],
                taxonomy,
            )
            if match and match.confidence >= 0.82:
                bucket = parent_buckets.setdefault(match.rule.name, {"rule": match.rule, "atomics": [], "matches": [], "v2": False})
                bucket["atomics"].append(atomic)
                bucket["matches"].append(match)
            elif atomic.category:
                fallback_buckets.setdefault(_key(atomic.category), []).append(atomic)
            else:
                unclassified.append(atomic)

    groups = []
    for bucket in parent_buckets.values():
        rule, atomics, matches = bucket["rule"], bucket["atomics"], bucket["matches"]
        common_reason = Counter(getattr(match, "matched_reason", getattr(match, "reason", "taxonomy match")) for match in matches).most_common(1)[0][0]
        weighted_confidence = 0.0
        subcategories = []
        evidence = []
        manual_review = False
        for match, atomic in zip(matches, atomics):
            if isinstance(match, TaxonomyClassification):
                weighted_confidence += (0.95 if match.confidence == "high" else 0.78 if match.confidence == "medium" else 0.55) * len(atomic.tickets)
                if match.subcategory:
                    subcategories.append(match.subcategory)
                evidence.extend(match.evidence)
                manual_review = manual_review or match.manual_review_recommended
            else:
                weighted_confidence += match.confidence * len(atomic.tickets)
        total_tickets = sum(len(atomic.tickets) for atomic in atomics)
        groups.append({
            "groupName": rule.name,
            "description": rule.description,
            "incidentCount": total_tickets,
            "representativeTickets": _representatives(atomics, representative_count),
            "confidence": round(weighted_confidence / total_tickets, 2) if total_tickets else 0.0,
            "matched_reason": common_reason,
            "subcategory": Counter(subcategories).most_common(1)[0][0] if subcategories else None,
            "evidence": evidence[:8],
            "manual_review_recommended": manual_review,
        })
    for category_key, atomics in fallback_buckets.items():
        category = next(atomic.category for atomic in atomics if atomic.category)
        groups.append({
            "groupName": f"{category.strip().title()} Issues",
            "description": f"Users are affected by recurring {category.strip().lower()} service problems.",
            "incidentCount": sum(len(atomic.tickets) for atomic in atomics),
            "representativeTickets": _representatives(atomics, representative_count),
            "confidence": 0.82,
            "matched_reason": f'structured category fallback: "{category}"',
        })
    return groups, unclassified


SEMANTIC_BOILERPLATE = (
    "user reported", "support team", "initial triage", "similar incidents", "business user",
    "ticket was reopened", "generic words", "service portal", "needs further investigation",
)
NAME_STOP_WORDS = STOP_WORDS | {
    "failed", "fails", "failure", "error", "errors", "not", "cannot", "unable", "prevents",
    "observed", "detected", "rejected", "requested", "request", "recurring", "same", "again",
    "during", "while", "because", "correctly", "working", "work", "longer", "expected",
    "delayed", "unavailable", "blocked", "requires", "review", "stalled",
    "regain", "sending", "sends", "does", "respond", "displayed", "changed", "visible",
}


def _semantic_core(ticket: NormalizedTicket) -> str:
    text = ticket.title or ticket.description or ticket.primary_text
    lowered = text.lower()
    marker_positions = [lowered.find(marker) for marker in SEMANTIC_BOILERPLATE if lowered.find(marker) > 0]
    if marker_positions:
        text = text[:min(marker_positions)].strip(" .;:-")
    sentences = [sentence.strip() for sentence in re.split(r"(?<=[.!?])\s+", text) if sentence.strip()]
    meaningful = [sentence for sentence in sentences if not any(marker in sentence.lower() for marker in SEMANTIC_BOILERPLATE)]
    return (meaningful[0] if meaningful else sentences[0] if sentences else text).strip()


def _discovered_label(atomics: list[_AtomicGroup]) -> str:
    cores = [_semantic_core(atomic.tickets[0]) for atomic in atomics]
    token_sets = [list(dict.fromkeys(re.findall(r"[a-z][a-z0-9-]{2,}", core.lower()))) for core in cores]
    frequency = Counter(token for tokens in token_sets for token in tokens if token not in NAME_STOP_WORDS)
    seed = token_sets[0] if token_sets else []
    selected = [token for token in seed if token not in NAME_STOP_WORDS and frequency[token] >= max(1, math.ceil(len(cores) * 0.35))]
    if len(selected) < 2:
        selected = [token for token, _ in frequency.most_common(4)]
    words = list(dict.fromkeys(selected))[:4]
    label = " ".join(words).strip().title() or "Emerging Service Pattern"
    if not re.search(r"\b(issue|failure|error|problem|alert)s?$", label, re.I):
        label += " Issues"
    return label


def _semantic_discovery_groups(
    atomics: list[_AtomicGroup],
    similarity_threshold: float,
    representative_count: int,
    min_group_size: int,
    taxonomy: tuple[TaxonomyRule, ...],
    embedding_method: str = "tfidf",
    clustering_method: str = "taxonomy_semantic",
    target_clusters: int | None = None,
    min_samples: int = 3,
) -> tuple[list[dict[str, Any]], list[_AtomicGroup]]:
    if not atomics:
        return [], []
    signatures = [_semantic_core(atomic.tickets[0]) for atomic in atomics]
    effective_threshold = min(0.72, max(0.42, similarity_threshold * 0.75))
    clusters, noise = _discovery_clusters(
        signatures,
        embedding_method=embedding_method,
        clustering_method=clustering_method,
        threshold=effective_threshold,
        target_clusters=target_clusters,
        min_samples=min_samples,
    )
    groups = []
    other_atomics: list[_AtomicGroup] = [atomics[index] for index in noise]
    for indices in clusters:
        members = [atomics[index] for index in indices]
        count = sum(len(atomic.tickets) for atomic in members)
        vectors = [_tokens(_semantic_core(atomic.tickets[0])) for atomic in members]
        similarities = [_cosine(vectors[left], vectors[right]) for left in range(len(vectors)) for right in range(left + 1, len(vectors))]
        cohesion = sum(similarities) / len(similarities) if similarities else 1.0
        confidence = min(0.9, 0.5 + min(0.18, math.log10(count + 1) * 0.07) + cohesion * 0.2)
        if count < min_group_size or confidence < 0.55:
            other_atomics.extend(members)
            continue
        label = _discovered_label(members)
        representative_text = [_semantic_core(atomic.tickets[0]) for atomic in members]
        v2_match = classify_ticket_v2({"summary": label, "description": " ".join(representative_text)}) if _using_default_taxonomy(taxonomy) else None
        taxonomy_match = find_taxonomy_match(None, label, representative_text, representative_text, taxonomy)
        if v2_match and v2_match.category != "Unclassified / Needs Review" and v2_match.score >= 4.5:
            label = v2_match.category
            description = v2_match.rule.description if v2_match.rule else f"A recurring pattern involving {_human_list(representative_text[:3])}."
            confidence = max(confidence, 0.78 if v2_match.confidence == "medium" else 0.9)
            reason = f'semantic merge: {v2_match.matched_reason}'
        elif taxonomy_match and taxonomy_match.confidence >= 0.9:
            label = taxonomy_match.rule.name
            description = taxonomy_match.rule.description
            confidence = max(confidence, 0.78)
            reason = f'semantic merge: {taxonomy_match.reason}'
        else:
            description = f"A recurring pattern involving {_human_list(representative_text[:3])}."
            reason = f"semantic discovery: {len(members)} issue signature(s), cohesion {cohesion:.2f}"
        groups.append({
            "groupName": label,
            "description": description,
            "incidentCount": count,
            "representativeTickets": _representatives(members, representative_count),
            "confidence": round(confidence, 2),
            "matched_reason": reason,
        })
    return groups, other_atomics


def _ticket_examples(atomic: _AtomicGroup, limit: int = 3) -> list[str]:
    examples = []
    for ticket in atomic.tickets[:limit]:
        text = ticket.title or ticket.description or ticket.primary_text
        if text:
            examples.append(text[:240])
    return examples


def _safe_float(value: Any, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_indices(value: Any, upper_bound: int, assigned: set[int]) -> list[int]:
    if not isinstance(value, list):
        return []
    indices = []
    for item in value:
        try:
            index = int(item)
        except (TypeError, ValueError):
            continue
        if 0 <= index < upper_bound and index not in assigned and index not in indices:
            indices.append(index)
    return indices


def _clean_suggestion(raw: dict[str, Any]) -> dict[str, Any] | None:
    name = str(raw.get("name") or raw.get("groupName") or "").strip()
    if not name:
        return None
    patterns = raw.get("patterns") or raw.get("suggestedIncludes") or raw.get("includes") or []
    contexts = raw.get("contexts") or []
    if not isinstance(patterns, list):
        patterns = []
    if not isinstance(contexts, list):
        contexts = []
    return {
        "name": name[:140],
        "description": str(raw.get("description") or "").strip()[:400],
        "patterns": [str(item).strip()[:80] for item in patterns if str(item).strip()][:12],
        "contexts": [str(item).strip()[:80] for item in contexts if str(item).strip()][:8],
        "reason": str(raw.get("reason") or raw.get("matchedReason") or "").strip()[:300],
    }


def _llm_assisted_unknown_groups(
    atomics: list[_AtomicGroup],
    representative_count: int,
    min_group_size: int,
    taxonomy: tuple[TaxonomyRule, ...],
    model: str | None,
    progress: Callable[[str, str], None] | None = None,
) -> tuple[list[dict[str, Any]], list[_AtomicGroup], list[dict[str, Any]], str]:
    if not atomics:
        return [], [], [], "not_needed"

    selected = atomics[:LLM_FALLBACK_MAX_ATOMICS]
    deferred = atomics[LLM_FALLBACK_MAX_ATOMICS:]
    selected_model = model or configured_model()
    unknown_payload = [
        {
            "index": index,
            "ticketCount": len(atomic.tickets),
            "category": atomic.category,
            "subcategory": atomic.subcategory,
            "examples": _ticket_examples(atomic),
        }
        for index, atomic in enumerate(selected)
    ]
    taxonomy_payload = [
        {
            "name": rule.name,
            "description": rule.description,
            "signals": list(rule.patterns[:8]),
        }
        for rule in taxonomy
    ]
    prompt = {
        "existingTaxonomy": taxonomy_payload,
        "unknownTicketClusters": unknown_payload,
        "instructions": {
            "grouping": "Group only clusters that share a real IT/business pain point. Reuse an existing taxonomy name when appropriate; otherwise create a concise parent group name.",
            "indices": "Use only the provided unknownTicketClusters index values. Do not invent indices.",
            "taxonomySuggestions": "Suggest durable taxonomy additions only for repeated or clearly reusable patterns.",
        },
    }
    try:
        diagnostic_event("ticket_analysis.llm_fallback.request", unknown_clusters=len(selected), model=selected_model)
        if progress:
            progress("gathering", f"Calling {selected_model} to classify {len(selected)} unknown ticket cluster(s) for taxonomy fallback")
        result = _json_object(_chat(
            "You are an ITSM taxonomy analyst. Return JSON only with keys: groups and taxonomySuggestions. "
            "groups is an array of {groupName, description, ticketIndices, confidence, matchedReason, suggestedIncludes}. "
            "taxonomySuggestions is an array of {name, description, patterns, contexts, reason}. "
            "Prefer stable parent pain points over narrow title rewrites. Leave unclear one-offs ungrouped.",
            json.dumps(prompt),
            selected_model,
            0.0,
            1800,
            max_retry_after_seconds=2,
        ))
    except Exception as exception:
        diagnostic_event("ticket_analysis.llm_fallback.error", exception_type=type(exception).__name__)
        if progress:
            progress("gathering", f"{selected_model} taxonomy fallback skipped: {type(exception).__name__}")
        return [], atomics, [], f"failed: {type(exception).__name__}"

    groups: list[dict[str, Any]] = []
    suggestions: list[dict[str, Any]] = []
    assigned: set[int] = set()
    raw_groups = result.get("groups") if isinstance(result, dict) else []
    if not isinstance(raw_groups, list):
        raw_groups = []
    for raw in raw_groups:
        if not isinstance(raw, dict):
            continue
        indices = _safe_indices(raw.get("ticketIndices") or raw.get("indices"), len(selected), assigned)
        if not indices:
            continue
        members = [selected[index] for index in indices]
        count = sum(len(atomic.tickets) for atomic in members)
        confidence = max(0.0, min(1.0, _safe_float(raw.get("confidence"), 0.64)))
        if confidence < LLM_FALLBACK_MIN_CONFIDENCE or (count < min_group_size and confidence < 0.76):
            continue
        name = str(raw.get("groupName") or "").strip()[:140]
        if not name:
            continue
        description = str(raw.get("description") or "").strip()[:500] or f"LLM-assisted grouping for {_human_list([_semantic_core(member.tickets[0]) for member in members[:3]])}."
        reason = str(raw.get("matchedReason") or "unknown tickets grouped by LLM fallback").strip()[:300]
        groups.append({
            "groupName": name,
            "description": description,
            "incidentCount": count,
            "representativeTickets": _representatives(members, representative_count),
            "confidence": round(confidence, 2),
            "matched_reason": f"llm fallback: {reason}",
        })
        assigned.update(indices)
        suggestion = _clean_suggestion(raw)
        if suggestion and suggestion["patterns"]:
            suggestions.append(suggestion)

    raw_suggestions = result.get("taxonomySuggestions") if isinstance(result, dict) else []
    if isinstance(raw_suggestions, list):
        for raw in raw_suggestions:
            if isinstance(raw, dict):
                suggestion = _clean_suggestion(raw)
                if suggestion and suggestion not in suggestions:
                    suggestions.append(suggestion)

    unresolved = [selected[index] for index in range(len(selected)) if index not in assigned] + deferred
    status = "used" if groups else "no_confident_groups"
    diagnostic_event("ticket_analysis.llm_fallback.response", groups=len(groups), unresolved=len(unresolved), suggestions=len(suggestions), status=status)
    if progress:
        progress("gathering", f"{selected_model} taxonomy fallback returned {len(groups)} group(s), {len(unresolved)} unresolved cluster(s)")
    return groups, unresolved, suggestions, status


def _llm_relabel_groups(
    groups: list[dict[str, Any]],
    taxonomy: tuple[TaxonomyRule, ...] = (),
    model: str | None = None,
    progress: Callable[[str, str], None] | None = None,
) -> tuple[int, str]:
    """Rewrite final group names/descriptions with the LLM, in place.

    Naming is strictly 1:1 and cosmetic: the model may reword a label but can
    never merge, split, reorder, or re-assign groups, so ticket membership stays
    exactly where the deterministic pipeline put it. A proposal is rejected if it
    is empty, degenerate, or would collide with another group's key — so the
    reported rename count is what actually changed, never the group total.

    Groups carrying a curated taxonomy rule name keep it: those labels are the
    reviewed vocabulary the taxonomy exists to enforce, so only generated and
    clustered labels are open to rewriting. When the taxonomy is paused there is
    no curated vocabulary to protect and every group becomes a candidate.
    """
    curated = {normalize_signal(rule.name) for rule in taxonomy}
    candidates = [
        (index, group) for index, group in enumerate(groups)
        if _key(group.get("groupName")) not in LLM_NAMING_PROTECTED_NAMES
        and normalize_signal(str(group.get("groupName") or "")) not in curated
    ][:LLM_NAMING_MAX_GROUPS]
    if not candidates:
        return 0, "not_needed"

    selected_model = model or configured_model()
    payload = [
        {
            "index": position,
            "currentName": str(group.get("groupName") or ""),
            "currentDescription": str(group.get("description") or ""),
            "ticketCount": int(group.get("incidentCount") or 0),
            "exampleTickets": [
                str(ticket.get("title"))[:240]
                for ticket in (group.get("representativeTickets") or [])[:LLM_NAMING_MAX_EXAMPLES]
                if ticket.get("title")
            ],
        }
        for position, (_, group) in enumerate(candidates)
    ]
    prompt = {
        "problemGroups": payload,
        "instructions": {
            "scope": "Improve only the name and description of each group. Never merge, split, drop, or re-order groups.",
            "indices": "Return one entry per provided index. Use only the given index values.",
            "naming": "Name the underlying business/IT pain point in 3-8 words, specific enough to distinguish it from the other groups. No ticket ids, no counts, no vendor-neutral filler like 'Various Issues'.",
            "description": "One sentence describing what the tickets in the group have in common and the impact on users.",
            "keep": "If the current name is already accurate and specific, return it unchanged.",
        },
    }
    try:
        diagnostic_event("ticket_analysis.llm_naming.request", groups=len(candidates), model=selected_model)
        if progress:
            progress("llm_labels", f"Calling {selected_model} to improve {len(candidates)} problem group name(s)")
        result = _json_object(_chat(
            "You are an ITSM reporting analyst writing problem-group labels for an executive summary. "
            "Return JSON only with key groups: an array of {index, name, description}. "
            "Keep every group distinct from the others and grounded in its example tickets.",
            json.dumps(prompt),
            selected_model,
            0.0,
            1800,
            max_retry_after_seconds=2,
        ))
    except Exception as exception:
        diagnostic_event("ticket_analysis.llm_naming.error", exception_type=type(exception).__name__)
        if progress:
            progress("llm_labels", f"{selected_model} group naming skipped: {type(exception).__name__}")
        return 0, f"failed: {type(exception).__name__}"

    raw_groups = result.get("groups") if isinstance(result, dict) else []
    if not isinstance(raw_groups, list):
        raw_groups = []
    taken = {_normalized_group_key(str(group.get("groupName") or "")) for group in groups}
    renamed = 0
    for raw in raw_groups:
        if not isinstance(raw, dict):
            continue
        try:
            position = int(raw.get("index"))
        except (TypeError, ValueError):
            continue
        if not 0 <= position < len(candidates):
            continue
        group = candidates[position][1]
        current_name = str(group.get("groupName") or "")
        name = re.sub(r"\s+", " ", str(raw.get("name") or raw.get("groupName") or "")).strip()[:140]
        description = re.sub(r"\s+", " ", str(raw.get("description") or "")).strip()[:500]
        current_key = _normalized_group_key(current_name)
        new_key = _normalized_group_key(name)
        name_ok = bool(name) and len(name) >= 3 and re.search(r"[a-z]", name, re.I) is not None
        if name_ok and new_key != current_key and new_key in taken:
            name_ok = False  # would collide with another group's label
        changed = False
        if name_ok and name != current_name:
            taken.discard(current_key)
            taken.add(new_key)
            group["llm_original_name"] = current_name
            group["groupName"] = name
            changed = True
        if description and description != str(group.get("description") or ""):
            group["description"] = description
            changed = True
        if changed:
            group["llm_named"] = True
            renamed += 1

    status = "used" if renamed else "no_changes"
    diagnostic_event("ticket_analysis.llm_naming.response", renamed=renamed, candidates=len(candidates), status=status)
    if progress:
        progress("llm_labels", f"{selected_model} rewrote {renamed} of {len(candidates)} problem group label(s)")
    return renamed, status


def _llm_taxonomy_suggestions(
    atomics: list[_AtomicGroup],
    taxonomy: tuple[TaxonomyRule, ...],
    model: str | None = None,
    progress: Callable[[str, str], None] | None = None,
) -> tuple[list[dict[str, Any]], str]:
    """Propose durable taxonomy rules for clusters nothing claimed.

    Independent of LLM fallback: this only ever returns suggestions for a human
    to approve, and never assigns a ticket or creates a group, so it is safe to
    run on its own. Suggestions without concrete patterns are dropped — a rule
    with no signals cannot be applied later.
    """
    if not atomics:
        return [], "not_needed"
    selected = atomics[:LLM_FALLBACK_MAX_ATOMICS]
    selected_model = model or configured_model()
    prompt = {
        "existingTaxonomy": [{"name": rule.name, "signals": list(rule.patterns[:8])} for rule in taxonomy],
        "unmatchedClusters": [
            {"index": index, "ticketCount": len(atomic.tickets), "examples": _ticket_examples(atomic)}
            for index, atomic in enumerate(selected)
        ],
        "instructions": {
            "scope": "Propose new taxonomy rules only for patterns that will recur. Ignore one-off tickets.",
            "duplicates": "Do not propose a rule that duplicates an existing taxonomy entry.",
            "patterns": "Every rule needs concrete keyword patterns that would match future tickets.",
        },
    }
    try:
        diagnostic_event("ticket_analysis.taxonomy_suggestions.request", clusters=len(selected), model=selected_model)
        if progress:
            progress("taxonomy_suggestions", f"Calling {selected_model} for taxonomy rule suggestions on {len(selected)} unmatched cluster(s)")
        result = _json_object(_chat(
            "You are an ITSM taxonomy curator. Return JSON only with key taxonomySuggestions: "
            "an array of {name, description, patterns, contexts, reason}. Suggest nothing when no pattern is durable.",
            json.dumps(prompt),
            selected_model,
            0.0,
            1200,
            max_retry_after_seconds=2,
        ))
    except Exception as exception:
        diagnostic_event("ticket_analysis.taxonomy_suggestions.error", exception_type=type(exception).__name__)
        if progress:
            progress("taxonomy_suggestions", f"{selected_model} taxonomy suggestions skipped: {type(exception).__name__}")
        return [], f"failed: {type(exception).__name__}"

    raw = result.get("taxonomySuggestions") if isinstance(result, dict) else []
    suggestions: list[dict[str, Any]] = []
    if isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            suggestion = _clean_suggestion(item)
            if suggestion and suggestion["patterns"] and suggestion not in suggestions:
                suggestions.append(suggestion)
    status = "used" if suggestions else "no_suggestions"
    diagnostic_event("ticket_analysis.taxonomy_suggestions.response", suggestions=len(suggestions), status=status)
    if progress:
        progress("taxonomy_suggestions", f"{selected_model} proposed {len(suggestions)} taxonomy rule(s)")
    return suggestions, status


def _normalized_group_key(name: str) -> str:
    words = normalize_signal(name).split()
    while words and words[-1] in {"issue", "issues"}:
        words.pop()
    return " ".join(words)


def _merge_group(target: dict[str, Any], source: dict[str, Any], representative_count: int) -> None:
    target["incidentCount"] += source["incidentCount"]
    target["confidence"] = max(target.get("confidence", 0.0), source.get("confidence", 0.0))
    seen = {(_key(ticket.get("ticketId")), _key(ticket.get("title"))) for ticket in target["representativeTickets"]}
    for ticket in source["representativeTickets"]:
        fingerprint = (_key(ticket.get("ticketId")), _key(ticket.get("title")))
        if fingerprint not in seen and len(target["representativeTickets"]) < representative_count:
            target["representativeTickets"].append(ticket)
            seen.add(fingerprint)


def _consolidate_groups(
    groups: list[dict[str, Any]],
    taxonomy: tuple[TaxonomyRule, ...],
    representative_count: int,
) -> list[dict[str, Any]]:
    """Merge duplicate labels and roll generated leaf groups into taxonomy parents.

    This second taxonomy pass is intentionally applied after semantic clustering:
    generated labels and their representatives can reveal a parent relationship
    that was not visible when the original row lacked structured fields.
    """
    exact: dict[str, dict[str, Any]] = {}
    for group in groups:
        key = _normalized_group_key(group["groupName"])
        if key in exact:
            _merge_group(exact[key], group, representative_count)
        else:
            exact[key] = {**group, "representativeTickets": list(group["representativeTickets"])}

    rules_by_name = {normalize_signal(rule.name): rule for rule in taxonomy}
    consolidated: dict[str, dict[str, Any]] = {}
    for group in exact.values():
        representative_text = [ticket.get("title") or "" for ticket in group["representativeTickets"]]
        v2_match = classify_ticket_v2({"summary": group["groupName"], "description": " ".join(representative_text)}) if _using_default_taxonomy(taxonomy) else None
        rule = rules_by_name.get(normalize_signal(group["groupName"])) or match_taxonomy_rule(
            None,
            group["groupName"],
            representative_text,
            representative_text,
            taxonomy,
        )
        if v2_match and v2_match.rule and v2_match.category != "Unclassified / Needs Review" and v2_match.score >= 4.5:
            group["groupName"] = v2_match.rule.name
            group["description"] = v2_match.rule.description
            group["confidence"] = max(group.get("confidence", 0.0), 0.9 if v2_match.confidence == "high" else 0.78)
            group["matched_reason"] = f'post-processing taxonomy merge: {v2_match.matched_reason}'
            group.setdefault("evidence", list(v2_match.evidence))
            group.setdefault("subcategory", v2_match.subcategory)
            group["manual_review_recommended"] = group.get("manual_review_recommended", False) or v2_match.manual_review_recommended
        elif rule:
            group["groupName"] = rule.name
            group["description"] = rule.description
            group["confidence"] = max(group.get("confidence", 0.0), 0.9)
            group["matched_reason"] = f'post-processing taxonomy merge: {group.get("matched_reason", "generated label or representatives matched")}'
        key = _normalized_group_key(group["groupName"])
        if key in consolidated:
            _merge_group(consolidated[key], group, representative_count)
        else:
            consolidated[key] = group
    return list(consolidated.values())


def analyze_ticket_rows(
    rows: list[dict[str, Any]],
    max_groups: int = 20,
    min_group_size: int = 3,
    similarity_threshold: float = 0.45,
    representative_count: int = 3,
    taxonomy: tuple[TaxonomyRule, ...] = DEFAULT_TAXONOMY,
    pause_okf_taxonomy: bool = False,
    strategy: str = "taxonomy_then_cluster",
    embedding_method: str = "tfidf",
    clustering_method: str = "taxonomy_semantic",
    target_clusters: int | None = None,
    hdbscan_min_samples: int = 3,
    llm_fallback: bool = False,
    llm_labels: bool = False,
    suggest_taxonomy_rules: bool = False,
    llm_model: str | None = None,
    progress: Callable[[str, str], None] | None = None,
) -> dict[str, Any]:
    tickets, empty, duplicates = clean_tickets(rows)
    total = len(tickets)
    normalized_strategy = {
        "okf_first": "taxonomy_then_cluster",
        "taxonomy_semantic": "taxonomy_then_cluster",
        "cluster_first": "cluster_only",
        "okf_only": "taxonomy_only",
    }.get(strategy, strategy)
    active_taxonomy: tuple[TaxonomyRule, ...] = () if pause_okf_taxonomy or normalized_strategy == "cluster_only" else taxonomy
    taxonomy_only = normalized_strategy == "taxonomy_only" and not pause_okf_taxonomy
    if pause_okf_taxonomy or normalized_strategy == "cluster_only":
        groups, unclassified = [], _initial_groups(tickets)
    else:
        groups, unclassified = _hierarchical_groups(tickets, representative_count, active_taxonomy)
    taxonomy_suggestions: list[dict[str, Any]] = []
    llm_fallback_status = "disabled" if not llm_fallback else "not_needed"
    suggestion_status = "disabled" if not suggest_taxonomy_rules else "not_needed"
    held_for_suggestions: list[_AtomicGroup] = []
    if unclassified and taxonomy_only:
        groups.append(_other_service_group(unclassified, representative_count, "taxonomy only: unmatched tickets held for review"))
    elif unclassified:
        discovered_groups, unresolved = _semantic_discovery_groups(
            unclassified, similarity_threshold, representative_count, min_group_size, active_taxonomy,
            embedding_method=embedding_method,
            clustering_method=clustering_method,
            target_clusters=target_clusters,
            min_samples=hdbscan_min_samples,
        )
        groups.extend(discovered_groups)
        if unresolved and llm_fallback:
            llm_groups, unresolved, suggestions, llm_fallback_status = _llm_assisted_unknown_groups(
                unresolved,
                representative_count,
                min_group_size,
                active_taxonomy,
                llm_model,
                progress,
            )
            groups.extend(llm_groups)
            taxonomy_suggestions.extend(suggestions)
        held_for_suggestions = list(unresolved)
        if unresolved:
            groups.append(_other_service_group(unresolved, representative_count))
    elif taxonomy_only:
        held_for_suggestions = list(unclassified)
    # Suggestions are advisory only, so they run on whatever stayed unmatched
    # regardless of whether LLM fallback was allowed to create groups from it.
    if suggest_taxonomy_rules and held_for_suggestions:
        suggestions, suggestion_status = _llm_taxonomy_suggestions(
            held_for_suggestions, active_taxonomy, llm_model, progress,
        )
        for suggestion in suggestions:
            if suggestion not in taxonomy_suggestions:
                taxonomy_suggestions.append(suggestion)
    groups = _consolidate_groups(groups, active_taxonomy, representative_count)
    groups.sort(key=lambda group: group["incidentCount"], reverse=True)
    if len(groups) > max_groups:
        overflow = [
            group for group in groups[max_groups - 1:]
            if group["groupName"] == "Other Service Issues"
            or group["incidentCount"] < min_group_size * 3
            or group.get("confidence", 0) < 0.6
        ]
        retained = [group for group in groups if group not in overflow]
        if overflow:
            overflow_tickets = [ticket for group in overflow for ticket in group["representativeTickets"]]
            retained.append({
                "groupName": "Other Service Issues",
                "description": "Small or low-confidence patterns that could not yet form a reliable business pain point.",
                "incidentCount": sum(group["incidentCount"] for group in overflow),
                "representativeTickets": overflow_tickets[:representative_count],
                "confidence": 0.4,
                "matched_reason": "fallback: presentation limit applied only to tiny or low-confidence clusters",
            })
        # max_groups is a soft presentation target. Coherent, high-confidence
        # domains are never discarded merely to satisfy the numeric limit.
        groups = retained
    for group in groups:
        group["percentage"] = round((group["incidentCount"] / total) * 100, 1) if total else 0.0
    groups.sort(key=lambda group: group["incidentCount"], reverse=True)
    # Naming runs last, on the surviving groups only, so the LLM sees the same
    # labels the user will and no tokens are spent on groups that get capped.
    groups_renamed = 0
    llm_label_status = "disabled"
    if llm_labels:
        groups_renamed, llm_label_status = _llm_relabel_groups(groups, active_taxonomy, llm_model, progress)
    manifest = {
            "totalRows": len(rows), "validTickets": total, "emptyTicketsRemoved": empty,
            "duplicatesRemoved": duplicates, "processedTickets": total,
            "problemGroups": len(groups), "coverageStatus": "complete",
            "taxonomyRules": len(active_taxonomy), "llmFallbackStatus": llm_fallback_status,
            "llmLabelStatus": llm_label_status, "llmGroupsRenamed": groups_renamed,
            "taxonomySuggestionStatus": suggestion_status,
            "embeddingMethod": embedding_method, "clusteringMethod": clustering_method,
    }
    if pause_okf_taxonomy:
        manifest["okfTaxonomyPaused"] = True
    return {
        "manifest": manifest,
        "groups": groups,
        "taxonomySuggestions": taxonomy_suggestions,
    }


def analyze_ticket_file(path: Path, **kwargs) -> dict[str, Any]:
    return analyze_ticket_rows(read_ticket_rows(path), **kwargs)


def ticket_analysis_markdown(result: dict[str, Any]) -> str:
    manifest, groups = result["manifest"], result["groups"]
    removed = manifest["emptyTicketsRemoved"] + manifest["duplicatesRemoved"]
    lines = [
        "# Ticket Analysis", "", "## Overview", "",
        f"- Total tickets uploaded: {manifest['totalRows']}",
        f"- Valid tickets analyzed: {manifest['validTickets']}",
        f"- Empty/duplicate tickets removed: {removed}",
        f"- Problem groups found: {manifest['problemGroups']}",
        f"- Taxonomy rules available: {manifest.get('taxonomyRules', len(DEFAULT_TAXONOMY))}",
        f"- LLM fallback: {manifest.get('llmFallbackStatus', 'disabled')}",
        f"- LLM naming: {manifest.get('llmLabelStatus', 'disabled')} ({manifest.get('llmGroupsRenamed', 0)} group(s) relabelled)", "",
        "## Problem Groups", "", "| Rank | Problem Group | Count | % |", "|---:|---|---:|---:|",
    ]
    lines.extend(f"| {rank} | {group['groupName']} | {group['incidentCount']} | {group['percentage']:.1f}% |" for rank, group in enumerate(groups, 1))
    lines.extend(["", "## Group Details", ""])
    for rank, group in enumerate(groups, 1):
        lines.extend([
            f"### {rank}. {group['groupName']}",
            f"**Count:** {group['incidentCount']} incidents  ",
            f"**Description:** {group['description']}  ",
            f"**Confidence:** {group.get('confidence', 0):.2f}  ",
            f"**Matched reason:** {group.get('matched_reason', 'not recorded')}  ",
            "**Representative tickets:**",
        ])
        for ticket in group["representativeTickets"]:
            prefix = f"{ticket['ticketId']} - " if ticket["ticketId"] else ""
            lines.append(f"- {prefix}{ticket['title']}")
        lines.append("")
    suggestions = result.get("taxonomySuggestions") or []
    if suggestions:
        lines.extend(["## Taxonomy Suggestions", ""])
        for suggestion in suggestions:
            patterns = ", ".join(suggestion.get("patterns") or []) or "No durable pattern suggested"
            reason = suggestion.get("reason") or "LLM fallback observed recurring unknown tickets"
            lines.extend([
                f"### {suggestion.get('name', 'Suggested Rule')}",
                f"**Description:** {suggestion.get('description') or 'No description provided'}  ",
                f"**Suggested patterns:** {patterns}  ",
                f"**Reason:** {reason}",
                "",
            ])
    lines.extend(["## Data Quality", "", f"Coverage: **{manifest['coverageStatus']}** ({manifest['processedTickets']} of {manifest['validTickets']} valid tickets processed)."])
    return "\n".join(lines)
