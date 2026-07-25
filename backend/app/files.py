import csv
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


TEXT_EXTENSIONS = {".txt", ".md", ".json", ".html", ".css", ".js", ".jsx", ".py"}
TABULAR_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xlsm"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".svg"}
SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | TABULAR_EXTENSIONS | {".pdf", ".docx"}
TEXT_STOP_WORDS = {
    "about", "after", "again", "been", "being", "could", "from", "have", "into", "that",
    "their", "there", "these", "this", "when", "where", "which", "with", "would", "were",
    "will", "your", "they", "them", "then", "than", "also", "only", "some", "such",
}


def _display(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _profile_table(filename: str, sheet_name: str, rows: Iterable[tuple]) -> str:
    iterator = iter(rows)
    header_row = None
    for row in iterator:
        if any(_display(value) for value in row):
            header_row = row
            break
    if header_row is None:
        return f"## Sheet: {sheet_name}\nNo populated rows."

    headers = []
    used = Counter()
    for index, value in enumerate(header_row):
        base = _display(value) or f"Column {index + 1}"
        used[base] += 1
        headers.append(base if used[base] == 1 else f"{base} ({used[base]})")

    counters = [Counter() for _ in headers]
    numeric = [{"count": 0, "sum": 0.0, "min": None, "max": None} for _ in headers]
    months = [Counter() for _ in headers]
    nonempty = [0 for _ in headers]
    text_terms = [Counter() for _ in headers]
    text_examples = [[] for _ in headers]
    samples = []
    row_count = 0

    for raw_row in iterator:
        values = list(raw_row[:len(headers)]) + [None] * max(0, len(headers) - len(raw_row))
        if not any(_display(value) for value in values):
            continue
        row_count += 1
        if len(samples) < 5:
            samples.append([_display(value) for value in values])
        for index, value in enumerate(values):
            text = _display(value)
            if not text:
                continue
            nonempty[index] += 1
            month = None
            if isinstance(value, (datetime, date)):
                month = value.strftime("%Y-%m")
            else:
                match = re.match(r"^(\d{4})[-/](\d{1,2})", text)
                if match:
                    month = f"{match.group(1)}-{int(match.group(2)):02d}"
            if month:
                months[index][month] += 1
            number = None
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                number = float(value)
            elif isinstance(value, str):
                try:
                    number = float(value.replace(",", ""))
                except ValueError:
                    pass
            if number is not None:
                stats = numeric[index]
                stats["count"] += 1
                stats["sum"] += number
                stats["min"] = number if stats["min"] is None else min(stats["min"], number)
                stats["max"] = number if stats["max"] is None else max(stats["max"], number)
            if text in counters[index] or len(counters[index]) < 5_000:
                counters[index][text] += 1
            if isinstance(value, str) and not text.replace(",", "").replace(".", "", 1).isdigit():
                for term in re.findall(r"[A-Za-z][A-Za-z0-9_-]{3,}", text.lower()):
                    if term not in TEXT_STOP_WORDS and (term in text_terms[index] or len(text_terms[index]) < 5_000):
                        text_terms[index][term] += 1
                if len(text) >= 12 and text not in text_examples[index] and len(text_examples[index]) < 5:
                    text_examples[index].append(text[:300])

    lines = [
        f"## Sheet: {sheet_name}",
        f"Rows analyzed: {row_count:,}",
        f"Columns ({len(headers)}): {', '.join(headers)}",
        "",
        "### Group-wise summaries",
    ]
    group_count = 0
    for index, header in enumerate(headers):
        distinct = len(counters[index])
        if not row_count or not distinct or distinct > min(250, max(25, row_count // 2)):
            continue
        top_values = counters[index].most_common(15)
        if not top_values or top_values[0][1] < 2:
            continue
        group_count += 1
        lines.extend([f"#### {header}", "| Group | Count | Share |", "|---|---:|---:|"])
        for value, count in top_values:
            safe_value = value.replace("|", "\\|")
            lines.append(f"| {safe_value} | {count:,} | {(count / row_count) * 100:.1f}% |")
        lines.append("")
    if not group_count:
        lines.append("No low-cardinality grouping columns were detected.\n")

    meaningful_text_columns = [index for index, examples in enumerate(text_examples) if examples]
    if meaningful_text_columns:
        lines.extend(["### Text evidence", "Recurring terms and examples are derived generically from all populated text fields.", ""])
        for index in meaningful_text_columns:
            common_terms = ", ".join(term for term, _ in text_terms[index].most_common(10)) or "No recurring terms"
            lines.extend([
                f"#### {headers[index]}",
                f"- Recurring terms: {common_terms}",
                "- Representative values:",
                *[f"  - {example}" for example in text_examples[index]],
                "",
            ])

    numeric_columns = [index for index, stats in enumerate(numeric) if stats["count"]]
    if numeric_columns:
        lines.extend(["### Numeric summaries", "| Column | Values | Average | Minimum | Maximum |", "|---|---:|---:|---:|---:|"])
        for index in numeric_columns:
            stats = numeric[index]
            lines.append(f"| {headers[index]} | {stats['count']:,} | {stats['sum'] / stats['count']:.2f} | {stats['min']:.2f} | {stats['max']:.2f} |")
        lines.append("")

    date_columns = [index for index, counter in enumerate(months) if counter]
    if date_columns:
        lines.append("### Time distributions")
        for index in date_columns:
            lines.extend([f"#### {headers[index]} by month", "| Month | Count |", "|---|---:|"])
            for month, count in sorted(months[index].items()):
                lines.append(f"| {month} | {count:,} |")
            lines.append("")

    completeness = [f"{headers[index]}: {(nonempty[index] / row_count) * 100:.1f}% populated" if row_count else f"{headers[index]}: empty" for index in range(len(headers))]
    lines.extend(["### Data completeness", *[f"- {item}" for item in completeness], ""])

    if samples:
        lines.extend(["### Sample records", "| " + " | ".join(headers) + " |", "|" + "|".join("---" for _ in headers) + "|"])
        for row in samples:
            lines.append("| " + " | ".join(value.replace("|", "\\|") for value in row) + " |")

    return "\n".join(lines)


def _spreadsheet_profile(filename: str, path: Path) -> str:
    extension = path.suffix.lower()
    sections = [f"# Deterministic tabular profile: {filename}", "Profile version: 3", "All counts and statistics below were computed from the complete dataset before LLM processing.", ""]
    if extension in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                sections.append(_profile_table(filename, sheet.title, sheet.iter_rows(values_only=True)))
        finally:
            workbook.close()
    else:
        delimiter = "\t" if extension == ".tsv" else ","
        with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            reader = csv.reader(handle, delimiter=delimiter)
            sections.append(_profile_table(filename, "Data", (tuple(row) for row in reader)))
    return "\n\n".join(sections)


def extract_text_from_path(filename: str, path: Path) -> str:
    extension = Path(filename).suffix.lower()
    if extension in TABULAR_EXTENSIONS:
        return _spreadsheet_profile(filename, path)
    if extension in TEXT_EXTENSIONS:
        return path.read_text(encoding="utf-8", errors="replace")
    if extension == ".pdf":
        return "\n\n".join(
            f"--- PAGE {index} ---\n{page.extract_text() or ''}"
            for index, page in enumerate(PdfReader(path).pages, start=1)
        )
    if extension == ".docx":
        return "\n".join(paragraph.text for paragraph in Document(path).paragraphs)
    return ""


def relevant_excerpt(text: str, terms: list[str], length: int = 420) -> tuple[int, str]:
    clean = re.sub(r"\s+", " ", text).strip()
    lowered = clean.lower()
    score = sum(lowered.count(term) for term in terms)
    positions = [lowered.find(term) for term in terms if lowered.find(term) >= 0]
    start = max(0, (min(positions) if positions else 0) - 80)
    excerpt = clean[start:start + length]
    if start:
        excerpt = "…" + excerpt
    if start + length < len(clean):
        excerpt += "…"
    return score, excerpt
